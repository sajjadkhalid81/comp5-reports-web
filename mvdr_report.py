"""
COMP5 MVDR (Master Vendor Document Register) Report Generator
Source: COMP5_MVDR__Log.xlsx — sheet MasterVDRLog (header row 3)
Statuses per COMP5-SPM-PE-PRC-00013 procedure.
Never touches report_core.py or mr_tbe_report.py.
"""
import pandas as pd
import warnings
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
TEAL="1B5E6B"; LT_TEAL="E0F2F4"; WHITE="FFFFFF"; BLACK="000000"; GREY="595959"
DK_GREEN="1F7A3C"; LT_GREEN="C6EFCE"; TAB_RED="C00000"; LT_RED="FFE0E0"
AMBER="FFF0CC"; DK_AMBER="7F6000"; ORANGE="C55A11"; LT_ORANGE="FCE4D6"
DK_BLUE="1F3864"; LT_BLUE="DEEAF1"
THIN=Side(border_style="thin",color="BFBFBF")

VEND_COLORS={"Not Submitted by Vendor":(LT_RED,TAB_RED),"Under Saipem Review":(LT_BLUE,DK_BLUE),
             "Pending with Vendor":(LT_ORANGE,ORANGE),"Code A — Saipem Accepted":(LT_GREEN,DK_GREEN),
             "Code I — For Information":("EDEDED",GREY)}
CPY_COLORS={"Not Submitted to CPY":(LT_RED,TAB_RED),"Under CPY Review":(LT_BLUE,DK_BLUE),
            "Under Consolidation":(AMBER,DK_AMBER),"Pending with Vendor":(LT_ORANGE,ORANGE),
            "Code A — CPY Approved":(LT_GREEN,DK_GREEN),"Code I — For Information":("EDEDED",GREY)}
NC_MAP={TEAL:LT_TEAL,DK_BLUE:LT_BLUE,TAB_RED:LT_RED,DK_AMBER:AMBER,
        DK_GREEN:LT_GREEN,ORANGE:LT_ORANGE,GREY:"EDEDED"}

# ── Helpers ──────────────────────────────────────────────────────────────────
def _b(): return Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

def _c(ws,r,c,v="",bold=False,fg=BLACK,bg=None,align="left",wrap=False,sz=9):
    cell=ws.cell(row=r,column=c,value=v)
    try: cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
    except Exception: pass
    try: cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    except Exception: pass
    try:
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
    except Exception: pass
    try: cell.border=_b()
    except Exception: pass
    return cell

def _h(ws,r,c,lbl,bg=TEAL,fg=WHITE,sz=9):
    return _c(ws,r,c,lbl,bold=True,fg=fg,bg=bg,align="center",wrap=True,sz=sz)

def _fmt(v):
    if v is None: return ""
    try:
        if pd.isna(v): return ""
    except Exception: pass
    if isinstance(v,(datetime,pd.Timestamp)):
        try: return v.strftime("%d-%b-%Y")
        except Exception: return ""
    s=str(v).strip()
    return "" if s in ["-","—","nan","NaT","None",""] else s

def is_val(v):
    return str(v).strip() not in ["","-","—","nan","None","NaT"] if v is not None else False

def code_start(v):
    return str(v).strip()[:1].upper() if v and is_val(v) else ""

# ── Read & categorise ────────────────────────────────────────────────────────
def _read_mvdr(raw_bytes):
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        df=pd.read_excel(BytesIO(raw_bytes),sheet_name="MasterVDRLog",header=2)
    df.columns=[str(c).strip() for c in df.columns]
    df=df[df["S/N"].notna()&df["Client Document No."].notna()].copy().reset_index(drop=True)

    vs_list=[]; cs_list=[]
    for _,row in df.iterrows():
        vend_in=row.get("Trans INCOMING No..4"); saip_out=row.get("Trans OUTGOING No..4")
        saip_code=row.get("Response Code.4"); cpy_out=row.get("Trans OUTGOING No")
        cpy_in=row.get("Trans INCOMING No..5"); cpy_code=row.get("Response Code.5")
        cpy_dt=pd.to_datetime(row.get("Date Rcvd.5"),errors="coerce")
        saip_dt=pd.to_datetime(row.get("Date Sent.4"),errors="coerce")
        sc=code_start(saip_code); cc=code_start(cpy_code)
        if not is_val(vend_in): vs="Not Submitted by Vendor"
        elif not is_val(saip_out): vs="Under Saipem Review"
        elif sc=="A": vs="Code A — Saipem Accepted"
        elif sc in ["B","C","D"]: vs="Pending with Vendor"
        elif sc=="I": vs="Code I — For Information"
        else: vs="Under Saipem Review"
        vs_list.append(vs)
        if not is_val(cpy_out): cs="Not Submitted to CPY"
        elif not is_val(cpy_in): cs="Under CPY Review"
        elif cc=="A": cs="Code A — CPY Approved"
        elif cc in ["B","C","D"]:
            if is_val(saip_out) and pd.notna(saip_dt) and pd.notna(cpy_dt) and saip_dt>=cpy_dt:
                cs="Pending with Vendor"
            else: cs="Under Consolidation"
        elif cc=="I": cs="Code I — For Information"
        else: cs="Under CPY Review"
        cs_list.append(cs)

    df["_VS"]=vs_list; df["_CS"]=cs_list
    df["_PO"]=df["P.O. Number"].astype(str).str.strip().str.replace(r"\.0$","",regex=True)
    df["_VND"]=df["Vendor Name"].astype(str).str.strip()
    df["_DSC"]=df["P.O.Descrption"].astype(str).str.strip()
    return df

# ── Summary for web UI ───────────────────────────────────────────────────────
def summarise_mvdr(raw_bytes):
    df=_read_mvdr(raw_bytes)
    vc=df["_VS"].value_counts(); cc=df["_CS"].value_counts()
    return {
        "total":            int(len(df)),
        "po_count":         int(df["_PO"].nunique()),
        "vend_not_sub":     int(vc.get("Not Submitted by Vendor",0)),
        "vend_under_rev":   int(vc.get("Under Saipem Review",0)),
        "vend_pending":     int(vc.get("Pending with Vendor",0)),
        "vend_code_a":      int(vc.get("Code A — Saipem Accepted",0)),
        "cpy_not_sub":      int(cc.get("Not Submitted to CPY",0)),
        "cpy_under_rev":    int(cc.get("Under CPY Review",0)),
        "cpy_consol":       int(cc.get("Under Consolidation",0)),
        "cpy_pending":      int(cc.get("Pending with Vendor",0)),
        "cpy_code_a":       int(cc.get("Code A — CPY Approved",0)),
    }

# ── KPI section builder ──────────────────────────────────────────────────────
def _make_kpi_section(ws,banner_row,label_row,count_row,banner_txt,kpis,col_pairs):
    first=col_pairs[0][0]; last=col_pairs[-1][1]
    ws.merge_cells(start_row=banner_row,start_column=first,end_row=banner_row,end_column=last)
    b=ws.cell(banner_row,first,banner_txt)
    b.font=Font(name="Arial",bold=True,size=11,color=WHITE)
    b.fill=PatternFill("solid",fgColor="0D4A55")
    b.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[banner_row].height=22
    for (cs,ce),(lbl,cnt,bg) in zip(col_pairs,kpis):
        lc_bg=NC_MAP.get(bg,LT_TEAL)
        ws.merge_cells(start_row=label_row,start_column=cs,end_row=label_row,end_column=ce)
        lc=ws.cell(label_row,cs,lbl)
        lc.font=Font(name="Arial",bold=True,size=10,color=WHITE)
        lc.fill=PatternFill("solid",fgColor=bg)
        lc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        lc.border=_b()
        for col in range(cs,ce+1):
            ws.cell(label_row,col).fill=PatternFill("solid",fgColor=bg)
        ws.merge_cells(start_row=count_row,start_column=cs,end_row=count_row,end_column=ce)
        nc=ws.cell(count_row,cs,cnt)
        nc.font=Font(name="Arial",bold=True,size=28,color=bg)
        nc.fill=PatternFill("solid",fgColor=lc_bg)
        nc.alignment=Alignment(horizontal="center",vertical="center")
        nc.border=_b()
        for col in range(cs,ce+1):
            ws.cell(count_row,col).fill=PatternFill("solid",fgColor=lc_bg)
    ws.row_dimensions[label_row].height=34
    ws.row_dimensions[count_row].height=46

# ── Data tab builder ─────────────────────────────────────────────────────────
def _build_data_tab(wb,tab_name,tab_color,df_tab,alt_bg,date_str):
    ws=wb.create_sheet(tab_name); ws.sheet_properties.tabColor=tab_color; ws.freeze_panes="A4"
    n=len(df_tab); NCOLS=17
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1"); t=ws["A1"]
    t.value=f"{tab_name.upper()}  |  COMP5 MVDR  |  {date_str}"
    t.font=Font(name="Arial",bold=True,size=11,color=WHITE)
    t.fill=PatternFill("solid",fgColor=tab_color)
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=22
    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2"); s=ws["A2"]
    s.value=f"Total Records: {n}  |  Date: {date_str}"
    s.font=Font(name="Arial",italic=True,size=9,color=GREY)
    s.alignment=Alignment(horizontal="center",vertical="center")
    s.fill=PatternFill(fill_type=None); ws.row_dimensions[2].height=16
    hdrs=["#","Client Doc No.","Saipem Doc No.","Vendor Doc No.","Document Title",
          "Doc Class","Current Rev.",
          "Vendor TR → Saipem","Date Rcvd","Saipem TR → Vendor","Date Returned","Saipem Code",
          "Saipem TR → CPY","Date to CPY","CPY TR → Saipem","Date from CPY","CPY Code"]
    col_w=[4,28,28,22,42,9,8,22,12,22,12,20,22,12,22,12,22]
    for ci,(h,w) in enumerate(zip(hdrs,col_w),1):
        _h(ws,3,ci,h,bg=tab_color)
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=30
    ws.auto_filter.ref=f"A3:{get_column_letter(NCOLS)}3"
    if df_tab.empty:
        ws.merge_cells(f"A4:{get_column_letter(NCOLS)}4")
        ws.cell(4,1,"No records").font=Font(italic=True,color=GREY,size=9); return
    AF=PatternFill("solid",fgColor=alt_bg); WF=PatternFill("solid",fgColor=WHITE)
    for ri,(_,row) in enumerate(df_tab.sort_values(["_VS","Client Document No."]).iterrows(),4):
        bg_fill=AF if ri%2==0 else WF
        vs=row.get("_VS",""); cs=row.get("_CS","")
        vs_bg,vs_fg=VEND_COLORS.get(vs,(WHITE,BLACK))
        cs_bg,cs_fg=CPY_COLORS.get(cs,(WHITE,BLACK))
        vals=[ri-3,_fmt(row.get("Client Document No.","")),_fmt(row.get("Saipem Document No.","")),
              _fmt(row.get("Vendor Document No.","")),_fmt(row.get("Document Title (All Caps)","")),
              _fmt(row.get("Doc. Class","")),_fmt(row.get("Current Rev.","")),
              _fmt(row.get("Trans INCOMING No..4","")),_fmt(row.get("Date Rcvd.4","")),
              _fmt(row.get("Trans OUTGOING No..4","")),_fmt(row.get("Date Sent.4","")),vs,
              _fmt(row.get("Trans OUTGOING No","")),_fmt(row.get("Date Sent.5","")),
              _fmt(row.get("Trans INCOMING No..5","")),_fmt(row.get("Date Rcvd.5","")),cs]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(row=ri,column=ci,value=val)
            if ci==12:
                cell.fill=PatternFill("solid",fgColor=vs_bg)
                cell.font=Font(name="Arial",bold=True,color=vs_fg,size=9)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            elif ci==17:
                cell.fill=PatternFill("solid",fgColor=cs_bg)
                cell.font=Font(name="Arial",bold=True,color=cs_fg,size=9)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            else:
                cell.fill=bg_fill
                cell.font=Font(name="Arial",size=9,color=BLACK)
                cell.alignment=Alignment(horizontal="center" if ci in [1,6,7] else "left",vertical="center")
            try: cell.border=_b()
            except Exception: pass
        ws.row_dimensions[ri].height=14

# ── PO-wise summary tab ──────────────────────────────────────────────────────
def _build_po_wise_tab(wb, df, date_str):
    ws = wb.create_sheet("PO WISE SUMMARY")
    ws.sheet_properties.tabColor = "0D4A55"

    col_widths = [13,13,13,13,13,13]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Title
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"COMP5 — MVDR PO WISE SUMMARY  |  {date_str}"
    t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill = PatternFill("solid", fgColor=TEAL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    s = ws["A2"]
    s.value = "Prepared by: Khalid Sajjad  |  CONFIDENTIAL"
    s.font = Font(name="Arial", italic=True, size=9, color=GREY)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    row = 4
    for (po, desc, vnd), grp in df.groupby(["_PO","_DSC","_VND"]):
        vc = grp["_VS"].value_counts()
        cc = grp["_CS"].value_counts()
        n  = len(grp)

        # PO banner
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        b = ws.cell(row, 1, f"PO {po}  —  {desc}  —  {vnd}   ({n} documents)")
        b.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        b.fill = PatternFill("solid", fgColor="0D4A55")
        b.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # Section 1 sub-banner
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        s1 = ws.cell(row, 1, "SECTION 1 — VENDOR / SAIPEM")
        s1.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        s1.fill = PatternFill("solid", fgColor=TEAL)
        s1.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

        sec1 = [("Not Submitted\nby Vendor", int(vc.get("Not Submitted by Vendor",0)), TAB_RED),
                ("Under SPM\nReview",        int(vc.get("Under Saipem Review",0)),      DK_BLUE),
                ("Pending\nwith Vendor",     int(vc.get("Pending with Vendor",0)),       ORANGE),
                ("Code A\nAccepted",         int(vc.get("Code A — Saipem Accepted",0)), DK_GREEN),
                ("Total\nDocuments",         n,                                          TEAL)]
        for ci, (lbl, cnt, bg) in enumerate(sec1, 1):
            lc = ws.cell(row, ci, lbl)
            lc.font = Font(name="Arial", bold=True, size=8, color=WHITE)
            lc.fill = PatternFill("solid", fgColor=bg)
            lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            lc.border = _b()
            nc = ws.cell(row+1, ci, cnt)
            nc.font = Font(name="Arial", bold=True, size=18, color=bg)
            nc.fill = PatternFill("solid", fgColor=NC_MAP.get(bg, LT_TEAL))
            nc.alignment = Alignment(horizontal="center", vertical="center")
            nc.border = _b()
        ws.row_dimensions[row].height = 26
        ws.row_dimensions[row+1].height = 30
        row += 2

        # Section 2 sub-banner
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        s2 = ws.cell(row, 1, "SECTION 2 — SAIPEM / CPY")
        s2.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        s2.fill = PatternFill("solid", fgColor=TEAL)
        s2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

        sec2 = [("Not Submitted\nto CPY",   int(cc.get("Not Submitted to CPY",0)),  TAB_RED),
                ("Under CPY\nReview",       int(cc.get("Under CPY Review",0)),       DK_BLUE),
                ("Under\nConsolidation",    int(cc.get("Under Consolidation",0)),   DK_AMBER),
                ("Pending\nwith Vendor",    int(cc.get("Pending with Vendor",0)),    ORANGE),
                ("Code A\nCPY Approved",    int(cc.get("Code A — CPY Approved",0)), DK_GREEN),
                ("Total\nDocuments",        n,                                       TEAL)]
        for ci, (lbl, cnt, bg) in enumerate(sec2, 1):
            lc = ws.cell(row, ci, lbl)
            lc.font = Font(name="Arial", bold=True, size=8, color=WHITE)
            lc.fill = PatternFill("solid", fgColor=bg)
            lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            lc.border = _b()
            nc = ws.cell(row+1, ci, cnt)
            nc.font = Font(name="Arial", bold=True, size=18, color=bg)
            nc.fill = PatternFill("solid", fgColor=NC_MAP.get(bg, LT_TEAL))
            nc.alignment = Alignment(horizontal="center", vertical="center")
            nc.border = _b()
        ws.row_dimensions[row].height = 26
        ws.row_dimensions[row+1].height = 30
        row += 3  # 2 rows + 1 blank between POs


# ── Main generator ───────────────────────────────────────────────────────────
def generate_mvdr(raw_bytes):
    df=_read_mvdr(raw_bytes)
    date_str=datetime.today().strftime("%d %B %Y")
    date_file=datetime.today().strftime("%d%b%Y").upper()
    vc=df["_VS"].value_counts(); cc=df["_CS"].value_counts(); total=len(df)

    wb=Workbook(); wb.remove(wb.active)
    ws=wb.create_sheet("SUMMARY"); ws.sheet_properties.tabColor=TEAL

    col_widths=[16,36,22,7,13,13,13,13,13,13,13,10]
    for ci,w in enumerate(col_widths,1):
        ws.column_dimensions[get_column_letter(ci)].width=w

    ws.merge_cells("A1:L1"); t=ws["A1"]
    t.value=f"COMP5 — MASTER VENDOR DOCUMENT REGISTER (MVDR)  |  {date_str}"
    t.font=Font(name="Arial",bold=True,size=14,color=WHITE)
    t.fill=PatternFill("solid",fgColor=TEAL)
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=36

    ws.merge_cells("A2:L2"); s=ws["A2"]
    s.value="Source: COMP5_MVDR__Log.xlsx  |  Prepared by: Khalid Sajjad  |  CONFIDENTIAL"
    s.font=Font(name="Arial",italic=True,size=9,color=GREY)
    s.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=16

    _make_kpi_section(ws,4,5,6,"SECTION 1 — VENDOR / SAIPEM TRANSMITTALS",
        [("NOT SUBMITTED\nby Vendor",int(vc.get("Not Submitted by Vendor",0)),TAB_RED),
         ("UNDER\nSaipem Review",int(vc.get("Under Saipem Review",0)),DK_BLUE),
         ("PENDING\nwith Vendor",int(vc.get("Pending with Vendor",0)),ORANGE),
         ("CODE A\nSaipem Accepted",int(vc.get("Code A — Saipem Accepted",0)),DK_GREEN),
         ("TOTAL\nDocuments",total,TEAL)],
        [(2,3),(4,5),(6,7),(8,9),(10,11)])

    _make_kpi_section(ws,8,9,10,"SECTION 2 — SAIPEM / CPY (QatarEnergy) STATUS",
        [("NOT SUBMITTED\nto CPY",int(cc.get("Not Submitted to CPY",0)),TAB_RED),
         ("UNDER\nCPY Review",int(cc.get("Under CPY Review",0)),DK_BLUE),
         ("UNDER\nConsolidation",int(cc.get("Under Consolidation",0)),DK_AMBER),
         ("PENDING\nwith Vendor",int(cc.get("Pending with Vendor",0)),ORANGE),
         ("CODE A\nCPY Approved",int(cc.get("Code A — CPY Approved",0)),DK_GREEN),
         ("TOTAL\nDocuments",total,TEAL)],
        [(1,2),(3,4),(5,6),(7,8),(9,10),(11,12)])

    ws.merge_cells("A12:L12"); pb=ws["A12"]
    pb.value="PO SUMMARY"
    pb.font=Font(name="Arial",bold=True,size=11,color=WHITE)
    pb.fill=PatternFill("solid",fgColor=TEAL)
    pb.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[12].height=20

    PO_HDRS=["PO Number","PO Description","Vendor","Total",
             "Not Sub\n(Vendor)","Under SPM\nReview","Pending\nVend (SPM)",
             "Not Sub\n(CPY)","Under CPY\nReview","Under\nConsol","Pending\nVend (CPY)","CPY\nCode A"]
    for ci,h in enumerate(PO_HDRS,1): _h(ws,13,ci,h,bg=TEAL)
    ws.row_dimensions[13].height=36

    ALT="E0F2F4"
    po_groups=list(df.groupby(["_PO","_DSC","_VND"]))
    for ri,((po,desc,vnd),grp) in enumerate(po_groups,14):
        bg=ALT if ri%2==0 else WHITE
        vc2=grp["_VS"].value_counts(); cc3=grp["_CS"].value_counts()
        vals=[po,desc,vnd,len(grp),
              int(vc2.get("Not Submitted by Vendor",0)),
              int(vc2.get("Under Saipem Review",0)),
              int(vc2.get("Pending with Vendor",0)),
              int(cc3.get("Not Submitted to CPY",0)),
              int(cc3.get("Under CPY Review",0)),
              int(cc3.get("Under Consolidation",0)),
              int(cc3.get("Pending with Vendor",0)),
              int(cc3.get("Code A — CPY Approved",0))]
        for ci,val in enumerate(vals,1):
            _c(ws,ri,ci,val,bg=bg,align="center" if ci>3 else "left")
        ws.row_dimensions[ri].height=16

    tr=14+len(po_groups)
    for col in range(1,13):
        cell=ws.cell(tr,col)
        try:
            cell.fill=PatternFill("solid",fgColor=TEAL)
            cell.font=Font(name="Arial",bold=True,size=9,color=WHITE)
            cell.alignment=Alignment(horizontal="center",vertical="center")
            cell.border=_b()
        except Exception: pass
    ws.cell(tr,1).value="TOTAL"; ws.cell(tr,4).value=total
    ws.cell(tr,5).value=int(vc.get("Not Submitted by Vendor",0))
    ws.cell(tr,6).value=int(vc.get("Under Saipem Review",0))
    ws.cell(tr,7).value=int(vc.get("Pending with Vendor",0))
    ws.cell(tr,8).value=int(cc.get("Not Submitted to CPY",0))
    ws.cell(tr,9).value=int(cc.get("Under CPY Review",0))
    ws.cell(tr,10).value=int(cc.get("Under Consolidation",0))
    ws.cell(tr,11).value=int(cc.get("Pending with Vendor",0))
    ws.cell(tr,12).value=int(cc.get("Code A — CPY Approved",0))
    ws.row_dimensions[tr].height=18

    _build_po_wise_tab(wb, df, date_str)

    _build_data_tab(wb,"ALL DOCUMENTS",      TEAL,    df,                                     LT_TEAL,  date_str)
    _build_data_tab(wb,"NOT SUBMITTED",      TAB_RED, df[df["_CS"]=="Not Submitted to CPY"],  LT_RED,   date_str)
    _build_data_tab(wb,"UNDER CPY REVIEW",   DK_BLUE, df[df["_CS"]=="Under CPY Review"],      LT_BLUE,  date_str)
    _build_data_tab(wb,"UNDER CONSOLIDATION",DK_AMBER,df[df["_CS"]=="Under Consolidation"],   AMBER,    date_str)
    _build_data_tab(wb,"PENDING WITH VENDOR",ORANGE,  df[df["_CS"]=="Pending with Vendor"],   LT_ORANGE,date_str)
    _build_data_tab(wb,"CODE A - APPROVED",  DK_GREEN,df[df["_CS"]=="Code A — CPY Approved"],LT_GREEN, date_str)

    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return {"bytes":buf.read(),"filename":f"COMP5_MVDR_Report_{date_file}.xlsx"}
