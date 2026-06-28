"""
mr_tbe_report.py  —  COMP5 QatarEnergy LNG / Saipem JV
Additional reports — completely separate from report_core.py.

Generates two Excel files from the Reference sheet of COMP5_REGISTER.xlsx:
  1. COMP5_MR_Report_<DATE>.xlsx   — H - MATERIAL REQUISITIONS
  2. COMP5_TBE_Report_<DATE>.xlsx  — J - TBE

Each file has 6 tabs (same concept as TQY/SDR):
  SUMMARY                  — 5 KPI boxes + discipline breakdown + expired urgent table
  ALL ISSUED               — All docs (latest rev per doc)
  NOT REPLIED (Not Expired)— Pending, within due date + Days Remaining
  NOT REPLIED (Expired)    — Overdue, past due date   + Days Overdue
  REPLIED CLOSED           — CPY responded with A (Approved)
  REPLIED OPEN             — CPY responded with B/C/D (comments/returned)

KPI boxes (5 total — no separate TOTAL box):
  1. ISSUED (to CPY)
  2. NOT REPLIED (Not Expired)
  3. NOT REPLIED (Expired)
  4. REPLIED CLOSED
  5. REPLIED OPEN

Discipline extracted from doc number 3rd segment:
  COMP5-SPM-PI-REQ-00001  →  PI
"""
from __future__ import annotations
from io import BytesIO
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette (mirrors report_core.py) ───────────────────────────────────────
WHITE     = "FFFFFF"
BLACK     = "000000"
DARK_BLUE = "1F3864"
GREY595   = "595959"
YELLOW    = "FFF2CC"
LT_RED    = "FFE0E0"
LT_GREEN  = "E2EFDA"
DK_GREEN  = "375623"
PURPLE    = "7030A0"
AMBER     = "FFF0CC"

MR_COLOR  = "1F6B75"   # teal
MR_ALT    = "D6EEF1"
TBE_COLOR = "4B2B7A"   # dark purple
TBE_ALT   = "E8DFFF"

TAB_GREEN  = "1F7A3C"
TAB_RED    = "C00000"
TAB_GREY   = GREY595
TAB_PURPLE = PURPLE

THIN = Side(border_style="thin", color="BFBFBF")

REV_ORDER = {
    "A1":1,  "A2":2,  "A3":3,  "A4":4,  "A5":5,
    "A":10, "B":11, "C":12, "D":13, "E":14, "F":15,
    "G":16, "H":17, "I":18, "J":19,
    "0":20, "00":20, "01":21, "02":22, "03":23, "04":24,
    "50":50, "51":51, "52":52,
}

DISC_MAP = {
    "AB":"Architectural",          "CC":"Construction",
    "CE":"Corrosion Engineering",  "CL":"Civil",
    "DR":"Drilling",               "EL":"Electrical",
    "GS":"Geosciences",            "HV":"HVAC",
    "IC":"ICS Security",           "IN":"Instrumentation",
    "LP":"HSE&Q/LOSPE",            "ME":"Mechanical",
    "MO":"Marine Operations",      "MT":"Material Technology",
    "OP":"Operations",             "PE":"Administrative / Eng. Mgmt",
    "PI":"Piping",                 "PL":"Pipelines",
    "PR":"Process",                "QM":"Quality Management",
    "SH":"HSE&Q/LOSPE",            "SS":"Subsurface",
    "ST":"Structural",             "TC":"Telecommunications",
}

DATA_COLS = [
    "#",
    "Document No. (with Rev)",
    "Document No.",
    "Rev",
    "Document Title",
    "Discipline",
    "Transmittal to CPY",
    "Date Issued to CPY",
    "Response Due Date",
    "Transmittal from CPY",
    "Date Replied",
    "CPY Response",
]
COL_WIDTHS = [5, 32, 28, 6, 42, 22, 26, 16, 16, 26, 16, 38]


# ── Helpers ────────────────────────────────────────────────────────────────

def _b():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _c(ws, r, c, v="", bold=False, fg=BLACK, bg=None,
        align="left", wrap=False, sz=9, brd=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if brd:
        cell.border = _b()
    return cell

def _h(ws, r, c, lbl, bg=DARK_BLUE, fg=WHITE, sz=9):
    return _c(ws, r, c, lbl, bold=True, fg=fg, bg=bg,
              align="center", wrap=True, sz=sz)

def _fmt(v):
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, float) and str(v) in ("nan","inf","-inf"):
        return ""
    if isinstance(v, (datetime, date, pd.Timestamp)):
        try:
            return v.strftime("%d-%b-%Y")
        except (ValueError, AttributeError):
            return ""
    s = str(v).strip()
    if s in ("–","-","—","nan","NaT","None"):
        return ""
    return s

def _wb_bytes(wb):
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


# ── Read & prepare ─────────────────────────────────────────────────────────

def _read_reference(raw: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(raw), sheet_name="Reference", header=0)
    df.columns = [str(c).strip() for c in df.columns]
    for col in ["DATE", "RESPONSE DUE DATE", "DATE REPLIED"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    # Extract discipline from 3rd segment of CLIENT DOCUMENT NO.
    doc_col = "CLIENT DOCUMENT NO."
    if doc_col in df.columns:
        df["_DISC"] = (
            df[doc_col].astype(str).str.split("-").str[2]
            .str.strip().str.upper()
        )
    else:
        df["_DISC"] = "XX"
    return df


def _dedup(df: pd.DataFrame) -> pd.DataFrame:
    """Keep latest revision per CLIENT DOCUMENT NO."""
    doc_col = "CLIENT DOCUMENT NO."
    rev_col = "REV"
    if doc_col not in df.columns or rev_col not in df.columns:
        return df
    df = df.copy()
    df["_REV_NUM"] = df[rev_col].astype(str).str.strip().map(REV_ORDER).fillna(99)
    df = df.sort_values([doc_col, "_REV_NUM"])
    df = df.drop_duplicates(subset=doc_col, keep="last")
    df = df.drop(columns=["_REV_NUM"])
    return df.reset_index(drop=True)


def _categorise(df: pd.DataFrame, today: date):
    """
    Split into 5 buckets — same logic as TQY/SDR:
      issued      — all (after dedup)
      not_rep_v   — no reply, due date >= today  (Days Remaining)
      not_rep_e   — no reply, due date <  today  (Days Overdue)
      rep_closed  — replied, response code starts with A
      rep_open    — replied, response code starts with B/C/D
    """
    df = df.copy()
    df["DATE REPLIED"]     = pd.to_datetime(df["DATE REPLIED"],     errors="coerce")
    df["RESPONSE DUE DATE"]= pd.to_datetime(df["RESPONSE DUE DATE"],errors="coerce")

    mask_replied = df["DATE REPLIED"].notna()
    resp_code    = df["STATUS / COMPANY RESPONSE"].astype(str).str.strip().str[:1].str.upper()
    due_dates    = df["RESPONSE DUE DATE"].dt.date

    issued    = df.copy()
    not_rep_v = df[~mask_replied & (due_dates >= today)].copy()
    not_rep_e = df[~mask_replied & (due_dates <  today)].copy()
    rep_closed = df[mask_replied & (resp_code == "A")].copy()
    rep_open   = df[mask_replied & (resp_code != "A")].copy()

    # Days Remaining (positive = days left)
    not_rep_v["Days Remaining"] = not_rep_v["RESPONSE DUE DATE"].apply(
        lambda d: (d.date() - today).days if pd.notna(d) else ""
    )
    # Days Overdue (positive = days past due)
    not_rep_e["Days Overdue"] = not_rep_e["RESPONSE DUE DATE"].apply(
        lambda d: (today - d.date()).days if pd.notna(d) else ""
    )

    return issued, not_rep_v, not_rep_e, rep_closed, rep_open


# ── Summary tab ────────────────────────────────────────────────────────────

def _build_summary(wb, title, header_color,
                   issued, not_rep_v, not_rep_e, rep_closed, rep_open,
                   report_date_str):
    """
    5 KPI boxes (no TOTAL box — issued count already shows total):
      1. ISSUED (to CPY)      — header_color
      2. NOT REPLIED (Not Exp)— green
      3. NOT REPLIED (Expired)— red
      4. REPLIED CLOSED       — grey
      5. REPLIED OPEN         — purple
    Then discipline breakdown table + TOTAL row + expired urgent table.
    """
    ws = wb.create_sheet("SUMMARY")
    ws.sheet_properties.tabColor = header_color

    # Row 1: Title banner
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=header_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Row 2: Report date
    ws.merge_cells("A2:F2")
    d = ws["A2"]
    d.value     = f"Report Date: {report_date_str}"
    d.font      = Font(name="Arial", italic=True, size=10, color=GREY595)
    d.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── 5 KPI boxes (row 4 = label, row 5 = count) ──
    nc_bg_map = {
        MR_COLOR:   "D6EEF1",
        TBE_COLOR:  "E8DFFF",
        TAB_GREEN:  "C6EFCE",
        TAB_RED:    "FFCCCC",
        TAB_GREY:   "E0E0E0",
        TAB_PURPLE: "EAD1FF",
    }
    kpis = [
        ("ISSUED\n(to CPY)",           len(issued),    header_color),
        ("NOT REPLIED\n(Not Expired)", len(not_rep_v), TAB_GREEN),
        ("NOT REPLIED\n(Expired)",     len(not_rep_e), TAB_RED),
        ("REPLIED\nCLOSED",            len(rep_closed), TAB_GREY),
        ("REPLIED\nOPEN",              len(rep_open),  TAB_PURPLE),
    ]
    for i, (label, count, bg) in enumerate(kpis, 1):
        lc = ws.cell(4, i, label)
        lc.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        lc.fill      = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lc.border    = _b()
        nc = ws.cell(5, i, count)
        nc.font      = Font(name="Arial", bold=True, size=28, color=bg)
        nc.fill      = PatternFill("solid", fgColor=nc_bg_map.get(bg, "EBF3FB"))
        nc.alignment = Alignment(horizontal="center", vertical="center")
        nc.border    = _b()

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.row_dimensions[4].height = 32
    ws.row_dimensions[5].height = 44

    # ── Discipline breakdown banner ──
    ws.merge_cells("A7:F7")
    banner = ws.cell(7, 1, "DISCIPLINE BREAKDOWN")
    banner.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    banner.fill      = PatternFill("solid", fgColor=header_color)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 20

    DISC_HDR_ROW = 8
    disc_headers = [
        "Discipline", "ISSUED\n(to CPY)", "NOT REPLIED\n(Not Exp.)",
        "NOT REPLIED\n(Expired)", "REPLIED\nCLOSED", "REPLIED\nOPEN"
    ]
    for ci, h in enumerate(disc_headers, 1):
        _h(ws, DISC_HDR_ROW, ci, h, bg=header_color)
    ws.row_dimensions[DISC_HDR_ROW].height = 30

    def get_disc(df_):
        return df_["_DISC"].astype(str).str.strip() if "_DISC" in df_.columns else pd.Series(dtype=str)

    all_discs = sorted(set(
        get_disc(issued).tolist()    + get_disc(not_rep_v).tolist() +
        get_disc(not_rep_e).tolist() + get_disc(rep_closed).tolist() +
        get_disc(rep_open).tolist()
    ) - {"", "nan", "NaN"})

    ALT = "EBF3FB"
    for ri, disc in enumerate(all_discs, DISC_HDR_ROW + 1):
        bg = ALT if ri % 2 == 0 else WHITE
        def cnt(df_, d=disc):
            s = get_disc(df_)
            return int((s == d).sum()) if len(s) else 0
        disc_name = DISC_MAP.get(disc, disc)
        exp_cnt   = cnt(not_rep_e)
        _c(ws, ri, 1, disc_name,       bg=bg)
        _c(ws, ri, 2, cnt(issued),     bg=bg, align="center")
        _c(ws, ri, 3, cnt(not_rep_v),  bg=bg, align="center")
        ec = _c(ws, ri, 4, exp_cnt,
                bg="FFCCCC" if exp_cnt > 0 else bg, align="center")
        if exp_cnt > 0:
            ec.font = Font(name="Arial", bold=True, color="C00000", size=9)
        _c(ws, ri, 5, cnt(rep_closed), bg=bg, align="center")
        op_cnt = cnt(rep_open)
        oc = _c(ws, ri, 6, op_cnt,
                bg="EAD1FF" if op_cnt > 0 else bg, align="center")
        if op_cnt > 0:
            oc.font = Font(name="Arial", bold=True, color=TAB_PURPLE, size=9)
        ws.row_dimensions[ri].height = 14

    # TOTAL row
    total_row = DISC_HDR_ROW + len(all_discs) + 1
    _c(ws, total_row, 1, "TOTAL",          bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 2, len(issued),      bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 3, len(not_rep_v),   bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 4, len(not_rep_e),
       bg="C00000" if len(not_rep_e) > 0 else header_color,
       bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 5, len(rep_closed),  bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 6, len(rep_open),    bg=header_color, bold=True, fg=WHITE, align="center")
    ws.row_dimensions[total_row].height = 18

    # ── Expired urgent table ──
    if len(not_rep_e) > 0:
        urg_row = total_row + 3
        ws.merge_cells(f"A{urg_row}:G{urg_row}")
        urg = ws.cell(urg_row, 1, "⚠  EXPIRED — URGENT ACTION REQUIRED")
        urg.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
        urg.fill      = PatternFill("solid", fgColor="C00000")
        urg.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[urg_row].height = 22

        for ci, h in enumerate(
            ["#", "Document No.", "Document Title", "Discipline",
             "Date Issued to CPY", "Due Date", "Days Overdue"], 1
        ):
            _h(ws, urg_row + 1, ci, h, bg="C00000")
        ws.column_dimensions[get_column_letter(3)].width = 45
        ws.row_dimensions[urg_row + 1].height = 20

        today_d = datetime.today().date()
        exp_sorted = not_rep_e.sort_values("RESPONSE DUE DATE")
        for ri2, (_, row_data) in enumerate(exp_sorted.iterrows(), urg_row + 2):
            due       = pd.to_datetime(row_data.get("RESPONSE DUE DATE"), errors="coerce")
            issued_dt = pd.to_datetime(row_data.get("DATE"), errors="coerce")
            days_over = row_data.get("Days Overdue", "")
            disc_name = DISC_MAP.get(str(row_data.get("_DISC","")).strip(), str(row_data.get("_DISC","")))

            _c(ws, ri2, 1, ri2 - (urg_row + 1),                              bg="FFCCCC", align="center", bold=True)
            _c(ws, ri2, 2, _fmt(row_data.get("CLIENT DOCUMENT NO.","")),      bg="FFCCCC")
            _c(ws, ri2, 3, _fmt(row_data.get("DOCUMENT TITLE","")),           bg="FFCCCC", wrap=True)
            _c(ws, ri2, 4, disc_name,                                         bg="FFCCCC", align="center")
            _c(ws, ri2, 5, _fmt(issued_dt),                                   bg="FFCCCC", align="center")
            _c(ws, ri2, 6, _fmt(due),                                         bg="FFCCCC", align="center")
            ov = _c(ws, ri2, 7,
                    f"OVERDUE {days_over}d" if days_over != "" else "",
                    bg="FFCCCC", align="center")
            ov.font = Font(name="Arial", bold=True, color="C00000", size=9)
            ws.row_dimensions[ri2].height = 14


# ── Data tab ───────────────────────────────────────────────────────────────

def _build_data_tab(wb, tab_name, tab_color, alt_bg, df, extra_col=None):
    """
    Data tab — same structure as TQY/SDR:
    Row 1: banner | Row 2: record count (no fill) | Row 3: headers | Row 4+: data
    extra_col: 'Days Remaining' or 'Days Overdue' appended as last column
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A4"

    report_date_str = datetime.today().strftime("%d %B %Y")
    n_records       = len(df)
    cols            = list(DATA_COLS) + ([extra_col] if extra_col and extra_col in df.columns else [])
    n_cols          = len(cols)

    # Row 1 — title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(1, 1, f"{tab_name.upper()}   |   COMP5 PROJECT   |   {report_date_str}")
    t.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=tab_color)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — record count (no fill)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sb = ws.cell(2, 1,
        f"Total Records: {n_records}   |   {tab_name}   |   Date: {report_date_str}")
    sb.font      = Font(name="Arial", italic=True, size=9, color=GREY595)
    sb.alignment = Alignment(horizontal="center", vertical="center")
    sb.fill      = PatternFill(fill_type=None)
    ws.row_dimensions[2].height = 16

    # Row 3 — headers
    widths = COL_WIDTHS + ([16] if extra_col and extra_col in df.columns else [])
    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        _h(ws, 3, ci, col, bg=tab_color)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 30
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"

    if df.empty:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
        ws.cell(4, 1, "No records for this category").font = \
            Font(italic=True, color=GREY595, size=9)
        return

    df_sorted = df.sort_values(
        ["_DISC", "CLIENT DOCUMENT NO."], na_position="last"
    ) if "_DISC" in df.columns else df

    for ri, (_, row_data) in enumerate(df_sorted.iterrows(), 4):
        bg        = alt_bg if ri % 2 == 0 else WHITE
        disc_code = str(row_data.get("_DISC","")).strip()
        disc_name = DISC_MAP.get(disc_code, disc_code)

        _c(ws, ri,  1, ri - 3,                                                   bg=bg, align="center")
        _c(ws, ri,  2, _fmt(row_data.get("CLIENT DOCUMENT NO. with REV","")),    bg=bg)
        _c(ws, ri,  3, _fmt(row_data.get("CLIENT DOCUMENT NO.","")),              bg=bg)
        _c(ws, ri,  4, _fmt(row_data.get("REV","")),                             bg=bg, align="center")
        _c(ws, ri,  5, _fmt(row_data.get("DOCUMENT TITLE","")),                  bg=bg)
        _c(ws, ri,  6, disc_name,                                                 bg=bg)
        _c(ws, ri,  7, _fmt(row_data.get("TRANSMITTAL #","")),                   bg=bg)
        _c(ws, ri,  8, _fmt(row_data.get("DATE","")),                            bg=bg, align="center")
        _c(ws, ri,  9, _fmt(row_data.get("RESPONSE DUE DATE","")),               bg=bg, align="center")
        _c(ws, ri, 10, _fmt(row_data.get("TRANSMITTAL FROM COMPANY","")),        bg=bg)
        _c(ws, ri, 11, _fmt(row_data.get("DATE REPLIED","")),                    bg=bg, align="center")
        _c(ws, ri, 12, _fmt(row_data.get("STATUS / COMPANY RESPONSE","")),       bg=bg)

        if extra_col and extra_col in df.columns:
            val = row_data.get(extra_col, "")
            if extra_col == "Days Remaining":
                try:
                    days = int(val)
                    disp = f"{days} days"
                    cell = _c(ws, ri, 13, disp,
                              bg=AMBER if days <= 7 else bg, align="center")
                    if days <= 7:
                        cell.font = Font(name="Arial", bold=True, color="7F6000", size=9)
                except (ValueError, TypeError):
                    _c(ws, ri, 13, _fmt(val), bg=bg, align="center")
            elif extra_col == "Days Overdue":
                try:
                    days = int(val)
                    cell = _c(ws, ri, 13, f"OVERDUE {days}d", bg="FFCCCC", align="center")
                    cell.font = Font(name="Arial", bold=True, color="C00000", size=9)
                except (ValueError, TypeError):
                    _c(ws, ri, 13, _fmt(val), bg=bg, align="center")

        ws.row_dimensions[ri].height = 14


# ── Main generate functions ─────────────────────────────────────────────────

def generate_mr(raw: bytes) -> dict:
    """MR — H - MATERIAL REQUISITIONS from Reference sheet."""
    today           = datetime.today().date()
    report_date_str = datetime.today().strftime("%d %B %Y")
    date_str        = datetime.today().strftime("%d%b%Y").upper()

    df_ref = _read_reference(raw)
    df_mr  = df_ref[df_ref["DOCUMENT TYPE"] == "H - MATERIAL REQUISITIONS"].copy()
    df_mr  = _dedup(df_mr)

    issued, not_rep_v, not_rep_e, rep_closed, rep_open = _categorise(df_mr, today)

    wb = Workbook(); wb.remove(wb.active)
    _build_summary(wb,
        title           = "MATERIAL REQUISITIONS (MR) | COMP5 PROJECT SUMMARY",
        header_color    = MR_COLOR,
        issued          = issued,
        not_rep_v       = not_rep_v,
        not_rep_e       = not_rep_e,
        rep_closed      = rep_closed,
        rep_open        = rep_open,
        report_date_str = report_date_str,
    )
    _build_data_tab(wb, "ALL ISSUED",               MR_COLOR,  MR_ALT,   issued)
    _build_data_tab(wb, "NOT REPLIED (Not Expired)", TAB_GREEN, "C6EFCE", not_rep_v, extra_col="Days Remaining")
    _build_data_tab(wb, "NOT REPLIED (Expired)",     TAB_RED,   "FFE0E0", not_rep_e, extra_col="Days Overdue")
    _build_data_tab(wb, "REPLIED CLOSED",            TAB_GREY,  "EDEDED", rep_closed)
    _build_data_tab(wb, "REPLIED OPEN",              TAB_PURPLE,"EAD1FF", rep_open)

    summary = {
        "issued":     len(issued),
        "not_rep_v":  len(not_rep_v),
        "not_rep_e":  len(not_rep_e),
        "rep_closed": len(rep_closed),
        "rep_open":   len(rep_open),
    }
    filename = f"COMP5_MR_Report_{date_str}.xlsx"
    return {"bytes": _wb_bytes(wb), "filename": filename, "summary": summary}


def generate_tbe(raw: bytes) -> dict:
    """TBE — J - TBE from Reference sheet."""
    today           = datetime.today().date()
    report_date_str = datetime.today().strftime("%d %B %Y")
    date_str        = datetime.today().strftime("%d%b%Y").upper()

    df_ref  = _read_reference(raw)
    df_tbe  = df_ref[df_ref["DOCUMENT TYPE"] == "J - TBE"].copy()
    df_tbe  = _dedup(df_tbe)

    issued, not_rep_v, not_rep_e, rep_closed, rep_open = _categorise(df_tbe, today)

    wb = Workbook(); wb.remove(wb.active)
    _build_summary(wb,
        title           = "TBE (TECHNICAL BID EVALUATION) | COMP5 PROJECT SUMMARY",
        header_color    = TBE_COLOR,
        issued          = issued,
        not_rep_v       = not_rep_v,
        not_rep_e       = not_rep_e,
        rep_closed      = rep_closed,
        rep_open        = rep_open,
        report_date_str = report_date_str,
    )
    _build_data_tab(wb, "ALL ISSUED",               TBE_COLOR, TBE_ALT,  issued)
    _build_data_tab(wb, "NOT REPLIED (Not Expired)", TAB_GREEN, "C6EFCE", not_rep_v, extra_col="Days Remaining")
    _build_data_tab(wb, "NOT REPLIED (Expired)",     TAB_RED,   "FFE0E0", not_rep_e, extra_col="Days Overdue")
    _build_data_tab(wb, "REPLIED CLOSED",            TAB_GREY,  "EDEDED", rep_closed)
    _build_data_tab(wb, "REPLIED OPEN",              TAB_PURPLE,"EAD1FF", rep_open)

    summary = {
        "issued":     len(issued),
        "not_rep_v":  len(not_rep_v),
        "not_rep_e":  len(not_rep_e),
        "rep_closed": len(rep_closed),
        "rep_open":   len(rep_open),
    }
    filename = f"COMP5_TBE_Report_{date_str}.xlsx"
    return {"bytes": _wb_bytes(wb), "filename": filename, "summary": summary}
