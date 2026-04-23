"""
report_core.py
==============
COMP5 — QatarEnergy LNG / Saipem JV
Weekly Document Control Report Generator

Generates three report types from the COMP5 project register:
  1. TQ  (Technical Query) Excel report        — via generate_tq_sdr()
  2. SDR (Specification Deviation Request) Excel report — via generate_tq_sdr()
  3. Weekly Issued Documents Excel report      — via generate_comp5()

Sources:
  - COMP5_Report_Instructions_v3.docx  (TQ & SDR specification)
  - COMP5_Report_Instructions.docx     (Issued Documents specification)

COMP5 Issued Documents — Key Rules:
  - Sheet: "Issued Documents" (header row 0)
  - LP discipline is merged into SH (LOSPE) before all processing
  - Only the latest revision per document number is counted
  - PCON- TR Issue Status categories:
      "Issued"               → forwarded to Company (CPY)
      "Not Issued"           → under process at PCON
      "Under Correction/Hold"→ pending with Engineering (NOT at PCON)
  - Status column shows: "Issued (0% Pending)" / "XX% Pending" /
    "Not Issued (XX% Pending)" — never "Partially Issued"
  - % Pending = (Not Issued + Under Correction/Hold) / Total × 100

Author : Document Control — COMP5
Version: 3.1.0
"""
from __future__ import annotations
from io import BytesIO
from datetime import datetime, date, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared palette ──────────────────────────────────────────────────────────
WHITE     = "FFFFFF"
BLACK     = "000000"
DARK_BLUE = "1F3864"
MID_BLUE  = "1F4E79"
LT_BLUE   = "DEEAF1"
DK_GREEN  = "375623"
LT_GREEN  = "E2EFDA"
DK_ORANGE = "C55A11"
MID_ORG   = "ED7D31"
LT_ORANGE = "FCE4D6"
YELLOW    = "FFF2CC"
LT_RED    = "FFE0E0"
PURPLE    = "7030A0"
LT_PURPLE = "EAD1FF"
LT_GREY   = "F2F2F2"
GREY595   = "595959"
AMBER     = "FFF0CC"

THIN = Side(border_style="thin",   color="BFBFBF")
MED  = Side(border_style="medium", color="595959")

def _b(med=False):
    s = MED if med else THIN
    return Border(left=s, right=s, top=s, bottom=s)

def _c(ws, r, c, v="", bold=False, fg=BLACK, bg=None,
        align="left", wrap=False, sz=9, brd=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if brd:
        cell.border = _b()
    return cell

def _h(ws, r, c, lbl, bg=DARK_BLUE, fg=WHITE, sz=9):
    return _c(ws, r, c, lbl, bold=True, fg=fg, bg=bg, align="center", wrap=True, sz=sz)

def _fmt(v):
    """Format a cell value. Returns empty string for None/NaT/NaN. Never returns dash."""
    if v is None:
        return ""
    # Must check pd.NaT BEFORE isinstance(datetime) — NaT passes that check
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, float) and str(v) in ("nan", "inf", "-inf"):
        return ""
    if isinstance(v, (datetime, date, pd.Timestamp)):
        try:
            return v.strftime("%d-%b-%Y")
        except (ValueError, AttributeError):
            return ""
    s = str(v).strip()
    # Never output bare dashes as placeholders
    if s in ("–", "-", "—", "nan", "NaT", "None"):
        return ""
    return s

def _wb_bytes(wb):
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


# ════════════════════════════════════════════════════════════════════════════
# SECTION 1 — TQ & SDR REPORT
# Source: COMP5_Report_Instructions_v3.docx
# ════════════════════════════════════════════════════════════════════════════

TQY_COLS = [
    "#", "Discipline", "Document Number", "Rev.", "Title",
    "Responsible Engineer", "Date Requested",
    "Transmittal to CPY", "Date Issued to CPY", "Due Date",
    "Transmittal from CPY", "Date Replied", "CPY Response", "Doc Status",
]
SDR_COLS = [
    "#", "Discipline", "Document Number", "Rev.", "Title",
    "Responsible Engineer", "Date Requested",
    "Transmittal to CPY", "Date Issued to CPY", "Due Date",
    "Transmittal from CPY", "Date Replied", "CPY Response",
]
# Source column names in the Excel (TQY sheet, header=2)
TQY_SRC_MAP = {
    "Discipline":          "Discipline",
    "Document Number":     "Document Number",
    "Rev.":                "Rev.",
    "Title":               "Title",
    "Responsible Engineer":"Responsible Engineer",
    "Date Requested":      "Date Requested",
    "Transmittal to CPY":  "TRANSMITTAL TO COMPANY",
    "Date Issued to CPY":  "DATE ISSUE TO CPY",
    "Due Date":            "RESPOND DUE DATE",
    "Transmittal from CPY":"TRANSMITTAL FROM COMPANY",
    "Date Replied":        "DATE REPLIED",
    "CPY Response":        "STATUS / COMPANY RESPONSE",
    "Doc Status":          "Document Status",
}
SDR_SRC_MAP = {
    "Discipline":          "Discipline",
    "Document Number":     "Document Number",
    "Rev.":                "Rev.",
    "Title":               "Title",
    "Responsible Engineer":"Responsible Engineer",
    "Date Requested":      "Date Requested",
    "Transmittal to CPY":  "TRANSMITTAL #",
    "Date Issued to CPY":  "DATE ISSUE TO CPY",
    "Due Date":            "RESPOND DUE DATE",
    "Transmittal from CPY":"TRANSMITTAL FROM COMPANY",
    "Date Replied":        "DATE REPLIED",
    "CPY Response":        "STATUS / COMPANY RESPONSE",
}

# Discipline code → full name (for Summary sheet display)
TQY_DISC_MAP = {
    "AB": "Architectural",
    "CC": "Construction",
    "CE": "Corrosion Engineering",
    "CL": "Civil",
    "DR": "Drilling",
    "EL": "Electrical",
    "GS": "Geosciences",
    "HV": "HVAC",
    "IC": "ICS Security",
    "IN": "Instrumentation",
    "LP": "HSE&Q/LOSPE",
    "ME": "Mechanical",
    "MO": "Marine Operations",
    "MT": "Material Technology",
    "OP": "Operations",
    "PE": "Administrative / Eng. Mgmt",
    "PI": "Piping",
    "PL": "Pipelines",
    "PR": "Process",
    "QM": "Quality Management",
    "SH": "HSE&Q/LOSPE",
    "SS": "Subsurface",
    "ST": "Structural",
    "TC": "Telecommunications",
}

# Tab colours per section 4
TAB_BLUE   = DARK_BLUE  # ISSUED
TAB_GREEN  = "1F7A3C"   # NOT REPLIED (Not Expired)
TAB_RED    = "C00000"   # NOT REPLIED (Expired)
TAB_GREY   = GREY595    # REPLIED CLOSED
TAB_PURPLE = PURPLE     # REPLIED OPEN
TAB_BROWN  = "7B3F00"   # SDR ISSUED


def _read_sheet(raw: bytes, sheet: str) -> pd.DataFrame:
    """Read TQY or SDR sheet. Header is at row index 2 (3rd row)."""
    df = pd.read_excel(BytesIO(raw), sheet_name=sheet, header=2)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=["Document Number"])
    df = df.reset_index(drop=True)
    # Parse all date-like columns
    for col in df.columns:
        if any(k in col.upper() for k in ("DATE", "DUE")):
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def categorise(df, transmittal_col, date_replied_col, response_col,
               rev_col, doc_col, today):
    """
    Exact categorisation logic from Section 7 of v3 instructions.
    Returns: issued, not_rep_valid, not_rep_expired, rep_closed, rep_open
    """
    df = df.copy()
    df[rev_col]          = df[rev_col].astype(str).str.strip().str.upper()
    df[date_replied_col] = pd.to_datetime(df[date_replied_col], errors="coerce")
    df["RESPOND DUE DATE"] = pd.to_datetime(df["RESPOND DUE DATE"], errors="coerce")

    # docs that appear more than once (multiple revisions exist)
    doc_rev_counts  = df.groupby(doc_col)[rev_col].count()
    docs_with_next  = set(doc_rev_counts[doc_rev_counts > 1].index)

    df["_resp_code"]   = df[response_col].astype(str).str.strip().str[:1].str.upper()
    df["_has_next_rev"]= df[doc_col].isin(docs_with_next)
    df["_replied"]     = df[date_replied_col].notna()

    # Keep LATEST revision only
    df_sorted = df.sort_values([doc_col, rev_col])
    df_latest = df_sorted.drop_duplicates(subset=[doc_col], keep="last").reset_index(drop=True)

    # Only include docs that have a real transmittal number (not blank/whitespace/nan/-)
    def _has_transmittal(series):
        """
        Only count if transmittal has a real value (not blank, space, NaN, dash).
        With PDC docs may have space-only values — exclude them.
        Not Received docs have NaN — exclude them.
        """
        s = series.fillna("").astype(str).str.strip()
        return s.notna() & (s != "") & (s != "nan") & (s != "NaN") & (s != "-") & (s != "–")

    issued = df_latest[_has_transmittal(df_latest[transmittal_col])]

    has_tr = _has_transmittal(df_latest[transmittal_col])
    not_rep_v = df_latest[
        has_tr &
        ~df_latest["_replied"] &
        (df_latest["RESPOND DUE DATE"] >= today)
    ]
    not_rep_e = df_latest[
        has_tr &
        ~df_latest["_replied"] &
        (df_latest["RESPOND DUE DATE"] < today)
    ]
    rep_closed = df_latest[
        df_latest["_replied"] &
        ((df_latest["_resp_code"] == "A") | df_latest["_has_next_rev"])
    ]
    rep_open = df_latest[
        df_latest["_replied"] &
        (df_latest["_resp_code"] != "A") &
        ~df_latest["_has_next_rev"]
    ]
    return issued, not_rep_v, not_rep_e, rep_closed, rep_open


def _map_cols(df: pd.DataFrame, src_map: dict, output_cols: list) -> pd.DataFrame:
    """Remap source columns to display column names. Add # as first col."""
    out = pd.DataFrame()
    for display_col in output_cols:
        if display_col == "#":
            continue
        src_col = src_map.get(display_col, display_col)
        if src_col in df.columns:
            out[display_col] = df[src_col].values
        else:
            out[display_col] = ""
    out.insert(0, "#", range(1, len(out)+1))
    return out


def _add_days_remaining(df: pd.DataFrame, src_df: pd.DataFrame, today) -> pd.DataFrame:
    """Add Days Remaining column. Values are signed integers."""
    due_col = "RESPOND DUE DATE"
    if due_col in src_df.columns:
        due = pd.to_datetime(src_df[due_col].values, errors="coerce")
        days = []
        for d in due:
            if pd.isna(d):
                days.append("")
            else:
                diff = (d.date() - today).days
                days.append(diff)
        df = df.copy()
        df["Days Remaining"] = days
    return df


def _build_tqsdr_summary(wb, title, header_color, kpis, disc_col,
                          issued, not_rep_v, not_rep_e, rep_closed, rep_open,
                          report_date_str):
    """Build Summary tab matching web report exactly:
    Row 1: Title banner (merged A:F)
    Row 2: Report date (merged A:F)
    Rows 4-9: 3 KPI boxes (label R4-6, count R7-9), each spanning 2 cols
    Rows 11-16: 2 KPI boxes (label R11-13, count R14-16)
    Row 18: DISCIPLINE BREAKDOWN banner
    Row 19: Disc table headers
    Row 20+: Disc data (NO total row)
    """
    ws = wb.create_sheet("SUMMARY")
    ws.sheet_properties.tabColor = header_color

    # Row 1: Title banner
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=header_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Row 2: Report date
    ws.merge_cells("A2:F2")
    d = ws["A2"]
    d.value     = f"Report Date: {report_date_str}"
    d.font      = Font(name="Arial", italic=True, size=10, color=GREY595)
    d.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # KPI boxes: 3 on top row (R4-9), 2 on bottom row (R11-16)
    # Each box: label rows span 3 rows, count rows span 3 rows
    # Columns: A:B, C:D, E:F
    BOX_POS = [(4,1),(4,2),(4,3),(4,4),(4,5)]
    for i, (label, count, fg, bg) in enumerate(kpis):
        row, col_idx = BOX_POS[i][0], BOX_POS[i][1]
        # Label cell (row 4)
        lc = ws.cell(row, col_idx, label)
        lc.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        lc.fill      = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lc.border    = _b()
        # Count cell (row 5)
        nc = ws.cell(row+1, col_idx, count)
        nc.font      = Font(name="Arial", bold=True, size=28, color=bg)
        nc_bg_map = {
            "1F4E79": "BDD7EE", "1F7A3C": "C6EFCE", "C00000": "FFCCCC",
            GREY595: "E0E0E0", PURPLE: "E2CFEF", "7B3F00": "FCE4D6",
        }
        nc.fill = PatternFill("solid", fgColor=nc_bg_map.get(bg, "EBF3FB"))
        nc.font = Font(name="Arial", bold=True, size=28, color=bg)
        nc.alignment = Alignment(horizontal="center", vertical="center")
        nc.border    = _b()

    # Col 6 must be blank — only 5 KPI boxes
    for rr in [4, 5]:
        c6 = ws.cell(rr, 6)
        c6.value = None
        c6.border = Border()
        c6.fill   = PatternFill(fill_type=None)

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.row_dimensions[4].height = 32
    ws.row_dimensions[5].height = 44

    # Row 18: Discipline breakdown banner
    DISC_HDR_ROW = 8
    ws.merge_cells(f"A7:F7")
    banner = ws.cell(7, 1, "DISCIPLINE BREAKDOWN")
    banner.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    banner.fill      = PatternFill("solid", fgColor=header_color)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 20

    # Row 19: Disc table column headers — "ISSUED" not "ISSUED (by Sai)"
    disc_headers = [
        "Discipline", "ISSUED\n(by Sai)", "NOT REPLIED\n(Not Exp.)",
        "NOT REPLIED\n(Expired)", "REPLIED\nCLOSED", "REPLIED\nOPEN"
    ]
    for ci, h in enumerate(disc_headers, 1):
        _h(ws, DISC_HDR_ROW, ci, h, bg=header_color)
    ws.row_dimensions[DISC_HDR_ROW].height = 30

    # Collect disciplines from ALL categories
    def get_disc(df_):
        if disc_col and disc_col in df_.columns:
            return df_[disc_col].astype(str).str.strip()
        return pd.Series(dtype=str)

    all_discs = sorted(set(
        get_disc(issued).tolist() +
        get_disc(not_rep_v).tolist() +
        get_disc(not_rep_e).tolist() +
        get_disc(rep_closed).tolist() +
        get_disc(rep_open).tolist()
    ) - {"", "nan", "NaN"})

    ALT = "EBF3FB"
    for ri, disc in enumerate(all_discs, DISC_HDR_ROW+1):
        bg = ALT if ri % 2 == 0 else WHITE
        def cnt(df_, d=disc):
            s = get_disc(df_)
            return int((s == d).sum()) if len(s) else 0
        disc_name = TQY_DISC_MAP.get(str(disc).strip().upper(), disc)
        _c(ws, ri, 1, disc_name, bg=bg)
        _c(ws, ri, 2, cnt(issued),     bg=bg, align="center")
        _c(ws, ri, 3, cnt(not_rep_v),  bg=bg, align="center")
        exp_cnt = cnt(not_rep_e)
        cell_exp = _c(ws, ri, 4, exp_cnt, bg="FFCCCC" if exp_cnt > 0 else bg, align="center")
        if exp_cnt > 0:
            cell_exp.font = Font(name="Arial", bold=True, color="C00000", size=9)
        _c(ws, ri, 5, cnt(rep_closed), bg=bg, align="center")
        _c(ws, ri, 6, cnt(rep_open),   bg=bg, align="center")
        ws.row_dimensions[ri].height = 14
    # TOTAL row
    total_row = DISC_HDR_ROW + len(all_discs) + 1
    def total_cnt(df_):
        if disc_col and disc_col in df_.columns:
            return len(df_)
        return 0
    _c(ws, total_row, 1, "TOTAL", bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 2, total_cnt(issued),    bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 3, total_cnt(not_rep_v), bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 4, total_cnt(not_rep_e), bg="C00000" if total_cnt(not_rep_e)>0 else header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 5, total_cnt(rep_closed),bg=header_color, bold=True, fg=WHITE, align="center")
    _c(ws, total_row, 6, total_cnt(rep_open),  bg=header_color, bold=True, fg=WHITE, align="center")
    ws.row_dimensions[total_row].height = 18

    # Expired urgent table (only if expired records exist)
    if len(not_rep_e) > 0 and disc_col and disc_col in not_rep_e.columns:
        urgent_row = DISC_HDR_ROW + len(all_discs) + 3
        ws.merge_cells(f"A{urgent_row}:G{urgent_row}")
        urg = ws.cell(urgent_row, 1, "⚠  EXPIRED — URGENT ACTION REQUIRED")
        urg.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
        urg.fill      = PatternFill("solid", fgColor="C00000")
        urg.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[urgent_row].height = 22
        for ci, h in enumerate(["#","Document Number","Discipline","Engineer","Date Issued to CPY","Due Date","Days Overdue"], 1):
            _h(ws, urgent_row+1, ci, h, bg="C00000")
        doc_col_name = "Document Number"
        ue_cols = ["Document Number","Discipline","RESPOND DUE DATE","DATE REPLIED","Responsible Engineer"]
        for ri2, (_, row_data) in enumerate(not_rep_e.iterrows(), urgent_row+2):
            due = pd.to_datetime(row_data.get("RESPOND DUE DATE"), errors="coerce")
            issued_date = pd.to_datetime(row_data.get("DATE ISSUE TO CPY"), errors="coerce")
            today_d = datetime.today().date()
            due_fmt = _fmt(due) if due is not None and not pd.isna(due) else ""
            issued_fmt = _fmt(issued_date) if issued_date is not None and not pd.isna(issued_date) else ""
            days_over = ""
            if due is not None and not pd.isna(due):
                try: days_over = (today_d - due.date()).days
                except (ValueError, AttributeError): pass
            _c(ws, ri2, 1, ri2-(urgent_row+1), bg="FFCCCC", align="center", bold=True)
            _c(ws, ri2, 2, _fmt(row_data.get("Document Number","")),  bg="FFCCCC")
            _c(ws, ri2, 3, _fmt(row_data.get("Discipline","")),        bg="FFCCCC", align="center")
            _c(ws, ri2, 4, _fmt(row_data.get("Responsible Engineer","")), bg="FFCCCC")
            _c(ws, ri2, 5, issued_fmt,                                 bg="FFCCCC", align="center")
            _c(ws, ri2, 6, due_fmt,                                    bg="FFCCCC", align="center")
            ov_cell = _c(ws, ri2, 7, f"OVERDUE {days_over}d" if days_over != "" else "", bg="FFCCCC", align="center")
            ov_cell.font = Font(name="Arial", bold=True, color="C00000", size=9)
            ws.row_dimensions[ri2].height = 14


def _build_tqsdr_data_tab(wb, tab_name, tab_color, src_df, src_map, out_cols,
                           is_not_replied=False):
    """
    Build a data tab.
    Row 1: TAB NAME  |  COMP5 PROJECT  |  DATE  (merged banner)
    Row 2: Total Records: N   |   TAB NAME   |   Date: DATE
    Row 3: Column headers
    Row 4+: Data (alternating colours)
    """
    today = datetime.today()
    report_date_str = today.strftime("%d %B %Y")

    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A4"

    display_df = _map_cols(src_df, src_map, out_cols)
    if is_not_replied:
        display_df = _add_days_remaining(display_df, src_df, today.date())

    n_cols = len(display_df.columns)
    n_records = len(display_df)

    # Row 1 — title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(1, 1, f"{tab_name.upper()}   |   COMP5 PROJECT   |   {report_date_str}")
    t.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=tab_color)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — record count bar (no background fill — match web report)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sb = ws.cell(2, 1, f"Total Records: {n_records}   |   {tab_name}   |   Date: {report_date_str}")
    sb.font      = Font(name="Arial", italic=True, size=9, color=GREY595)
    sb.alignment = Alignment(horizontal="center", vertical="center")
    sb.fill      = PatternFill(fill_type=None)  # no fill
    ws.row_dimensions[2].height = 16

    # Row 3 — column headers
    for ci, col in enumerate(display_df.columns, 1):
        _h(ws, 3, ci, col, bg=tab_color)
        # Set column width
        if col == "#":
            ws.column_dimensions[get_column_letter(ci)].width = 5
        elif col in ("Rev.", "Discipline"):
            ws.column_dimensions[get_column_letter(ci)].width = 10
        elif "Date" in col or col == "Due Date":
            ws.column_dimensions[get_column_letter(ci)].width = 15
        elif col in ("Document Number","Transmittal to CPY","Transmittal from CPY"):
            ws.column_dimensions[get_column_letter(ci)].width = 28
        elif col == "Title":
            ws.column_dimensions[get_column_letter(ci)].width = 40
        elif col == "Days Remaining":
            ws.column_dimensions[get_column_letter(ci)].width = 16
        else:
            ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.row_dimensions[3].height = 30

    # Row 4+ — data
    # Alternating colour per tab
    ALT_COLORS = {
        TAB_BLUE:   "DEEAF1",
        "1F7A3C":   "C6EFCE",
        "C00000":   "FFE0E0",
        GREY595:    "EDEDED",
        PURPLE:     "EAD1FF",
        TAB_BROWN:  "FCE4D6",
    }
    alt_bg = ALT_COLORS.get(tab_color, "EBF3FB")
    today_d = datetime.today().date()

    if display_df.empty:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
        ws.cell(4, 1, "No records for this category").font = Font(italic=True, color=GREY595, size=9)
        return

    dr_idx = list(display_df.columns).index("Days Remaining") + 1 if "Days Remaining" in display_df.columns else None

    for ri, (_, row_data) in enumerate(display_df.iterrows(), 4):
        bg = alt_bg if ri % 2 == 0 else WHITE
        for ci, col in enumerate(display_df.columns, 1):
            val = row_data[col]
            disp = _fmt(val)

            if col == "Days Remaining" and val != "":
                try:
                    days = int(val)
                    if days < 0:
                        disp = f"OVERDUE {abs(days)}d"
                        cell = _c(ws, ri, ci, disp, bg="FFCCCC", align="center")
                        cell.font = Font(name="Arial", bold=True, color="C00000", size=9)
                    elif days <= 7:
                        disp = f"{days} days"
                        cell = _c(ws, ri, ci, disp, bg=YELLOW, align="center")
                        cell.font = Font(name="Arial", bold=True, color="7F6000", size=9)
                    else:
                        disp = f"{days} days"
                        cell = _c(ws, ri, ci, disp, bg=bg, align="center")
                    continue
                except (ValueError, TypeError):
                    pass
            _c(ws, ri, ci, disp, bg=bg, align="center" if col == "#" else "left")
        ws.row_dimensions[ri].height = 14

    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"


def generate_tq_sdr(raw: bytes) -> dict:
    """
    Generate TQY and SDR Excel reports from the TQ-SDR log file.
    Implements COMP5_Report_Instructions_v3.docx exactly.
    """
    today        = pd.Timestamp.today().normalize()
    report_date  = datetime.today().strftime("%d %B %Y")
    date_str     = datetime.today().strftime("%d%b%Y").upper()

    # ── TQY ──────────────────────────────────────────────────────────────
    tqy = _read_sheet(raw, "TQY")
    disc_col_tqy = "Discipline" if "Discipline" in tqy.columns else None

    tqy_issued, tqy_valid, tqy_expired, tqy_closed, tqy_repopen = categorise(
        tqy,
        transmittal_col  = "TRANSMITTAL TO COMPANY",
        date_replied_col = "DATE REPLIED",
        response_col     = "STATUS / COMPANY RESPONSE",
        rev_col          = "Rev.",
        doc_col          = "Document Number",
        today            = today,
    )

    tqy_kpis = [
        ("TQ ISSUED\n(by Sai)",           len(tqy_issued),  WHITE,  DARK_BLUE),
        ("TQ NOT REPLIED\n(Not Expired)", len(tqy_valid),   WHITE,  "1F7A3C"),
        ("TQ NOT REPLIED\n(Expired)",     len(tqy_expired), WHITE,  "C00000"),
        ("TQ REPLIED\nCLOSED",            len(tqy_closed),  WHITE,  GREY595),
        ("TQ REPLIED\nOPEN",              len(tqy_repopen), WHITE,  PURPLE),
    ]
    wb_tqy = Workbook(); wb_tqy.remove(wb_tqy.active)
    _build_tqsdr_summary(wb_tqy, "TQ – TECHNICAL QUERY REGISTER | COMP5 PROJECT SUMMARY", TAB_BLUE, tqy_kpis,
                          disc_col_tqy,
                          tqy_issued, tqy_valid, tqy_expired, tqy_closed, tqy_repopen,
                          report_date)
    _build_tqsdr_data_tab(wb_tqy, "TQ ISSUED (by Sai)",           TAB_BLUE,   tqy_issued,  TQY_SRC_MAP, TQY_COLS)
    _build_tqsdr_data_tab(wb_tqy, "TQ NOT REPLIED (Not Expired)", "1F7A3C",   tqy_valid,   TQY_SRC_MAP, TQY_COLS, is_not_replied=True)
    _build_tqsdr_data_tab(wb_tqy, "TQ NOT REPLIED (Expired)",     "C00000",   tqy_expired, TQY_SRC_MAP, TQY_COLS, is_not_replied=True)
    _build_tqsdr_data_tab(wb_tqy, "TQ REPLIED CLOSED",            GREY595,    tqy_closed,  TQY_SRC_MAP, TQY_COLS)
    _build_tqsdr_data_tab(wb_tqy, "TQ REPLIED OPEN",              PURPLE,     tqy_repopen, TQY_SRC_MAP, TQY_COLS)

    # ── SDR ──────────────────────────────────────────────────────────────
    sdr = _read_sheet(raw, "SDR")
    disc_col_sdr = "Discipline" if "Discipline" in sdr.columns else None

    sdr_issued, sdr_valid, sdr_expired, sdr_closed, sdr_repopen = categorise(
        sdr,
        transmittal_col  = "TRANSMITTAL #",
        date_replied_col = "DATE REPLIED",
        response_col     = "STATUS / COMPANY RESPONSE",
        rev_col          = "Rev.",
        doc_col          = "Document Number",
        today            = today,
    )

    sdr_kpis = [
        ("SDR ISSUED\n(by Sai)",           len(sdr_issued),  WHITE,  TAB_BROWN),
        ("SDR NOT REPLIED\n(Not Expired)", len(sdr_valid),   WHITE,  "1F7A3C"),
        ("SDR NOT REPLIED\n(Expired)",     len(sdr_expired), WHITE,  "C00000"),
        ("SDR REPLIED\nCLOSED",            len(sdr_closed),  WHITE,  GREY595),
        ("SDR REPLIED\nOPEN",              len(sdr_repopen), WHITE,  PURPLE),
    ]
    wb_sdr = Workbook(); wb_sdr.remove(wb_sdr.active)
    _build_tqsdr_summary(wb_sdr, "SDR – SPECIFICATION DEVIATION REQUEST | COMP5 PROJECT SUMMARY", TAB_BROWN, sdr_kpis,
                          disc_col_sdr,
                          sdr_issued, sdr_valid, sdr_expired, sdr_closed, sdr_repopen,
                          report_date)
    _build_tqsdr_data_tab(wb_sdr, "SDR ISSUED (by Sai)",           TAB_BROWN, sdr_issued,  SDR_SRC_MAP, SDR_COLS)
    _build_tqsdr_data_tab(wb_sdr, "SDR NOT REPLIED (Not Expired)", "1F7A3C",  sdr_valid,   SDR_SRC_MAP, SDR_COLS, is_not_replied=True)
    _build_tqsdr_data_tab(wb_sdr, "SDR NOT REPLIED (Expired)",     "C00000",  sdr_expired, SDR_SRC_MAP, SDR_COLS, is_not_replied=True)
    _build_tqsdr_data_tab(wb_sdr, "SDR REPLIED CLOSED",            GREY595,   sdr_closed,  SDR_SRC_MAP, SDR_COLS)
    _build_tqsdr_data_tab(wb_sdr, "SDR REPLIED OPEN",              PURPLE,    sdr_repopen, SDR_SRC_MAP, SDR_COLS)

    summary = {
        "tqy": {"issued": len(tqy_issued), "valid": len(tqy_valid),
                "expired": len(tqy_expired), "closed": len(tqy_closed), "repopen": len(tqy_repopen)},
        "sdr": {"issued": len(sdr_issued), "valid": len(sdr_valid),
                "expired": len(sdr_expired), "closed": len(sdr_closed), "repopen": len(sdr_repopen)},
    }
    return {
        "tqy_bytes": _wb_bytes(wb_tqy), "tqy_name": f"COMP5_TQY_Report_{date_str}.xlsx",
        "sdr_bytes": _wb_bytes(wb_sdr), "sdr_name": f"COMP5_SDR_Report_{date_str}.xlsx",
        "summary":   summary,
    }


# ════════════════════════════════════════════════════════════════════════════
# SECTION 2 — WEEKLY COMP5 ISSUED DOCUMENTS REPORT
# Source: COMP5_Report_Instructions.docx
# ════════════════════════════════════════════════════════════════════════════

DISC_MAP = {
    "PR": "Process",        "EL": "Electrical",
    "SH": "LOSPE",          "ME": "Mechanical",
    "PI": "Piping",         "IN": "Instrumentation",
    "ST": "Structural",     "AB": "Architectural",
    "TC": "Telecommunication","CE": "Corrosion Engineering",
    "HV": "HVAC",           "PE": "Administrative / Eng. Mgmt",
    "LP": "HSE&Q/LOSPE",          # LP → SH/LOSPE
}
DISC_ORDER = ["PR","EL","SH","ME","PI","ST","TC","AB","CE","HV","IN","PE"]

REV_ORDER = {"00":0,"0":0,"A":1,"B":2,"C":3,"D":4,"E":5,"F":6,"G":7,"H":8,"I":9,"J":10}

DETAIL_COLS = [
    "#", "CLIENT DOCUMENT NO. with REV", "Saipem Number", "Rev",
    "TITLE / DESCRIPTION", "Disc.", "Discipline Name",
    "Issuing Description", "Date Issued", "Transmittal Reference",
    "Issued by DC", "PCON TR", "PCON- TR Issue Status",
]


def _read_comp5(raw: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(raw), sheet_name="Issued Documents", header=0)
    df.columns = [str(c).strip() for c in df.columns]
    # LP → SH
    if "Discipline" in df.columns:
        df["Discipline"] = df["Discipline"].astype(str).str.strip().str.upper()
        df["Discipline"] = df["Discipline"].replace("LP", "SH")
    # Parse dates
    for col in df.columns:
        if "date" in col.lower():
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def _dedup_comp5(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only latest revision per CLIENTDOCUMENTNO."""
    doc_col = "CLIENTDOCUMENTNO."
    rev_col = "Rev"
    if doc_col not in df.columns:
        # try to find it
        matches = [c for c in df.columns if "CLIENTDOC" in c.upper().replace(" ","") or "CLIENT DOC" in c.upper()]
        if matches:
            doc_col = matches[0]
        else:
            return df  # cannot dedup

    df = df.copy()
    df["Rev_num"] = df[rev_col].astype(str).str.strip().map(REV_ORDER).fillna(99) if rev_col in df.columns else 0
    sort_cols = [doc_col, "Rev_num"]
    if "Date Issued" in df.columns:
        sort_cols.append("Date Issued")
    df = df.sort_values(sort_cols)
    df = df.drop_duplicates(subset=doc_col, keep="last")
    df = df.drop(columns=["Rev_num"], errors="ignore")
    return df.reset_index(drop=True)


def _split_comp5(df: pd.DataFrame):
    """Split into 3 categories based on PCON- TR Issue Status."""
    status_col = "PCON- TR Issue Status"
    if status_col not in df.columns:
        # try to find it
        matches = [c for c in df.columns if "PCON" in c.upper() and "STATUS" in c.upper()]
        status_col = matches[0] if matches else None

    if status_col:
        s = df[status_col].astype(str).str.strip()
        df_issued    = df[s.str.upper() == "ISSUED"].copy()
        df_not_issued = df[s.str.upper() == "NOT ISSUED"].copy()
        df_hold      = df[s.str.upper().str.contains("CORRECTION|HOLD", na=False)].copy()
        df_all       = df.copy()
    else:
        df_issued = df_not_issued = df_hold = pd.DataFrame()
        df_all = df.copy()

    return df_all, df_issued, df_not_issued, df_hold


def _get_disc_code(disc_value):
    """Return uppercase disc code, treating LP as SH."""
    code = str(disc_value).strip().upper()
    return "SH" if code == "LP" else code


def _build_comp5_summary(wb, df, df_issued, df_not_issued, df_hold, report_date_str):
    """
    Summary tab: 7-column discipline breakdown table + legend.
    Columns: # | Code | Discipline Name | Issued to PDC | Under Process at PCON |
             Pending with Engineering | PCON TR Issue Status (% Pending)
    """
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = DARK_BLUE

    # Title banner
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = f"COMP5 — WEEKLY ISSUED DOCUMENTS STATUS REPORT  |  {report_date_str}"
    t.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Sub-title
    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = f"Cut-Off Date: {report_date_str}  |  Total Documents: see table below  |  CONFIDENTIAL"
    s.font      = Font(name="Arial", italic=True, size=9, color=GREY595)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Column headers row 4
    HDR_ROW = 4
    hdrs = ["#", "Code", "Discipline Name", "Issued\nto PDC",
            "Under Process\nat PCON", "Pending with\nEngineering",
            "PCON TR Issue Status\n(% Pending)"]
    col_widths = [5, 7, 28, 14, 18, 22, 24]
    for ci, (h, w) in enumerate(zip(hdrs, col_widths), 1):
        _h(ws, HDR_ROW, ci, h, bg=DARK_BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 32

    # Discipline rows
    disc_col = "Discipline"
    status_col = "PCON- TR Issue Status"

    def count_by_disc(df_, disc_code):
        if df_.empty or disc_col not in df_.columns:
            return 0
        return (df_[disc_col].astype(str).str.strip().str.upper() == disc_code).sum()

    ALT = "DEEAF1"
    for ri, code in enumerate(DISC_ORDER, HDR_ROW+1):
        bg = ALT if ri % 2 == 0 else WHITE
        disc_name = DISC_MAP.get(code, code)

        # Count all docs for this discipline (in full deduped df)
        total = count_by_disc(df, code)
        under = count_by_disc(df_not_issued, code)
        hold  = count_by_disc(df_hold, code)
        pending = under + hold
        issued_to_cpy = count_by_disc(df_issued, code)

        # Status cell logic
        if total == 0:
            status_txt = "No Docs"
            status_bg  = LT_GREY
            status_fg  = "666666"
        elif pending == 0:
            status_txt = "Issued (0% Pending)"
            status_bg  = LT_GREEN
            status_fg  = DK_GREEN
        elif issued_to_cpy == 0:
            pct = int(round(pending / total * 100))
            status_txt = f"Not Issued ({pct}% Pending)"
            status_bg  = LT_RED
            status_fg  = "C00000"
        else:
            pct = int(round(pending / total * 100))
            status_txt = f"{pct}% Pending"
            status_bg  = YELLOW
            status_fg  = "7F6000"

        _c(ws, ri, 1, ri - HDR_ROW, bg=bg, align="center")
        _c(ws, ri, 2, code, bg=bg, align="center", bold=True)
        _c(ws, ri, 3, disc_name, bg=bg)
        # Issued to PDC = total docs for this disc
        tc = _c(ws, ri, 4, total, bg=bg, align="center", bold=True)
        tc.font = Font(name="Arial", bold=True, color=MID_BLUE, size=9)
        # Under Process
        uc = _c(ws, ri, 5, under, bg=bg, align="center", bold=True)
        if under > 0:
            uc.font = Font(name="Arial", bold=True, color="C00000", size=9)
        # Pending with Engineering
        hc = _c(ws, ri, 6, hold, bg=AMBER if hold > 0 else bg, align="center", bold=(hold > 0))
        if hold > 0:
            hc.font = Font(name="Arial", bold=True, color="7F6000", size=9)
        # Status
        sc = _c(ws, ri, 7, status_txt, bg=status_bg, align="center", bold=True)
        sc.font = Font(name="Arial", bold=True, color=status_fg, size=9)
        ws.row_dimensions[ri].height = 16

    # ── TOTAL row ────────────────────────────────────────────────────────────
    total_row = HDR_ROW + len(DISC_ORDER) + 1
    ws.merge_cells(f"A{total_row}:C{total_row}")
    tc = ws.cell(total_row, 1, "TOTAL")
    tc.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    tc.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = _b()
    total_d2p    = sum(count_by_disc(df, code) for code in DISC_ORDER)
    total_under  = sum(count_by_disc(df_not_issued, code) for code in DISC_ORDER)
    total_hold   = sum(count_by_disc(df_hold, code) for code in DISC_ORDER)
    total_pend   = total_under + total_hold
    tot_pct      = int(round(total_pend / total_d2p * 100)) if total_d2p > 0 else 0
    for ci2, val in [(4, total_d2p), (5, total_under), (6, total_hold)]:
        tc2 = ws.cell(total_row, ci2, val)
        tc2.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        tc2.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        tc2.alignment = Alignment(horizontal="center", vertical="center")
        tc2.border    = _b()
    # Status cell in total row
    tc7 = ws.cell(total_row, 7, f"{tot_pct}% Pending")
    tc7.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    tc7.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    tc7.alignment = Alignment(horizontal="center", vertical="center")
    tc7.border    = _b()
    ws.row_dimensions[total_row].height = 20

    # ── Footnote ─────────────────────────────────────────────────────────────
    note_row = total_row + 1
    ws.merge_cells(f"A{note_row}:G{note_row}")
    nc = ws.cell(note_row, 1,
                 "* LP (Layout & Plot Plan) documents combined under SH – LOSPE.  "
                 "** Under Correction/Hold = pending with Engineering, NOT at PCON.")
    nc.font      = Font(name="Arial", italic=True, size=8, color="444444")
    nc.fill      = PatternFill("solid", fgColor=LT_GREY)
    nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    nc.border    = _b()
    ws.row_dimensions[note_row].height = 14

    # Legend
    leg_row = HDR_ROW + len(DISC_ORDER) + 4
    ws.merge_cells(f"A{leg_row}:G{leg_row}")
    lg = ws.cell(leg_row, 1, "LEGEND")
    lg.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    lg.fill = PatternFill("solid", fgColor=DARK_BLUE)
    lg.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[leg_row].height = 18

    legend = [
        (LT_GREEN,  DK_GREEN, "Issued (0% Pending) — All documents forwarded to Company"),
        (YELLOW,    "7F6000", "XX% Pending — Partially forwarded"),
        (LT_RED,    "C00000", "Not Issued (XX% Pending) — No documents forwarded yet"),
        (AMBER,     "7F6000", "Pending with Engineering — Under Correction/Hold (back with Eng.)"),
        (LT_GREY,   "666666", "No Docs — No documents for this discipline"),
    ]
    for ri2, (bg, fg, txt) in enumerate(legend, leg_row+1):
        ws.merge_cells(f"A{ri2}:G{ri2}")
        lc = ws.cell(ri2, 1, txt)
        lc.font      = Font(name="Arial", size=9, color=fg)
        lc.fill      = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        lc.border    = _b()
        ws.row_dimensions[ri2].height = 15


def _build_datewise_tab(wb, df, report_date_str):
    """
    Date-Wise Breakdown tab — ORANGE theme throughout.
    Section 1: docs per day by discipline.
    Section 2: daily stats with % columns.
    """
    ws = wb.create_sheet("Date-Wise Breakdown")
    ws.sheet_properties.tabColor = DK_ORANGE

    # Title banner — dynamic width based on disciplines + 2 (Date + TOTAL)
    # We don't know discs yet so use a safe wide range; will tighten after disc calc
    _TITLE_END_COL = get_column_letter(len(DISC_ORDER) + 2)
    ws.merge_cells(f"A1:{_TITLE_END_COL}1")
    t = ws["A1"]
    t.value     = f"COMP5 — DATE-WISE ISSUED DOCUMENTS BREAKDOWN  |  {report_date_str}"
    t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DK_ORANGE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    if "Date Issued" not in df.columns or df.empty:
        ws.cell(3, 1, "No date data available").font = Font(italic=True, color=GREY595, size=9)
        return

    date_col   = "Date Issued"
    disc_col   = "Discipline"
    status_col = "PCON- TR Issue Status"

    df2 = df.copy()
    df2["_date"] = pd.to_datetime(df2[date_col], errors="coerce").dt.date
    df2["_disc"] = df2[disc_col].astype(str).str.strip().str.upper().replace("LP","SH") if disc_col in df2.columns else "?"

    dates_sorted = sorted(df2["_date"].dropna().unique())
    discs_sorted = [c for c in DISC_ORDER if c in df2["_disc"].values]

    # ── SECTION 1: per-day by discipline ─────────────────────────────────
    sec1_hdr_row = 3
    _S1_END_COL = get_column_letter(len(discs_sorted) + 2)
    ws.merge_cells(f"A{sec1_hdr_row}:{_S1_END_COL}{sec1_hdr_row}")
    s1 = ws.cell(sec1_hdr_row, 1, "SECTION 1 — DOCUMENTS ISSUED PER DAY BY DISCIPLINE")
    s1.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    s1.fill      = PatternFill("solid", fgColor=MID_ORG)
    s1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sec1_hdr_row].height = 18

    n_discs = len(discs_sorted)
    hdr_cols = ["Date"] + discs_sorted + ["TOTAL"]
    hdr_row = sec1_hdr_row + 1
    for ci, h in enumerate(hdr_cols, 1):
        _h(ws, hdr_row, ci, h, bg=DK_ORANGE)
        ws.column_dimensions[get_column_letter(ci)].width = 12 if ci == 1 else 8
    ws.row_dimensions[hdr_row].height = 22

    row = hdr_row + 1
    all_total = 0
    disc_totals = {d: 0 for d in discs_sorted}
    for dt in dates_sorted:
        day_df = df2[df2["_date"] == dt]
        bg = LT_ORANGE if row % 2 == 0 else WHITE
        _c(ws, row, 1, dt.strftime("%d-%b-%Y") if dt else "", bg=bg, align="center")
        row_total = 0
        for ci2, disc in enumerate(discs_sorted, 2):
            cnt = (day_df["_disc"] == disc).sum()
            disc_totals[disc] += cnt
            row_total += cnt
            cell = _c(ws, row, ci2, cnt if cnt > 0 else "", bg="FFDAB9" if cnt > 0 else bg, align="center")
        all_total += row_total
        _c(ws, row, len(hdr_cols), row_total, bg=DK_ORANGE if row_total>0 else bg,
           align="center", bold=True)
        if row_total > 0:
            ws.cell(row, len(hdr_cols)).font = Font(name="Arial", bold=True, color=WHITE, size=9)
        ws.row_dimensions[row].height = 14
        row += 1

    # Total row
    _c(ws, row, 1, "TOTAL", bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    for ci2, disc in enumerate(discs_sorted, 2):
        tc = _c(ws, row, ci2, disc_totals[disc], bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    _c(ws, row, len(hdr_cols), all_total, bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    ws.row_dimensions[row].height = 18
    row += 2

    # ── SECTION 2: daily stats with % ─────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    s2 = ws.cell(row, 1, "SECTION 2 — DAILY STATUS SUMMARY")
    s2.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    s2.fill      = PatternFill("solid", fgColor=DK_ORANGE)
    s2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    sec2_hdrs = ["Date","Issued to PDC","Under Process\nat PCON","Pending with\nEngineering","% Under Process","% Issued to CPY"]
    for ci, h in enumerate(sec2_hdrs, 1):
        _h(ws, row, ci, h, bg=DK_ORANGE if ci in (1,6) else MID_ORG)
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.row_dimensions[row].height = 28
    row += 1

    for dt in dates_sorted:
        day_df = df2[df2["_date"] == dt]
        total  = len(day_df)
        if status_col in df2.columns:
            s = day_df[status_col].astype(str).str.strip().str.upper()
            issued_cpy = (s == "ISSUED").sum()
            not_iss    = (s == "NOT ISSUED").sum()
            hold       = s.str.contains("CORRECTION|HOLD", na=False).sum()
        else:
            issued_cpy = not_iss = hold = 0

        pct_under  = f"{not_iss / total * 100:.1f}%" if total > 0 else "0.0%"
        pct_issued = f"{issued_cpy / total * 100:.1f}%" if total > 0 else "0.0%"

        pct_u_val  = not_iss / total * 100 if total > 0 else 0
        pct_c_val  = issued_cpy / total * 100 if total > 0 else 0

        # % Under Process colour
        if pct_u_val == 0:    pu_bg, pu_fg = LT_GREEN, DK_GREEN
        elif pct_u_val <= 50: pu_bg, pu_fg = YELLOW,   "7F6000"
        else:                 pu_bg, pu_fg = LT_RED,    "C00000"

        # % Issued to CPY colour
        if pct_c_val == 100:  pc_bg, pc_fg = LT_GREEN, DK_GREEN
        elif pct_c_val > 0:   pc_bg, pc_fg = YELLOW,   "7F6000"
        else:                 pc_bg, pc_fg = LT_RED,    "C00000"

        bg = LT_ORANGE if row % 2 == 0 else WHITE
        _c(ws, row, 1, dt.strftime("%d-%b-%Y") if dt else "", bg=bg, align="center")
        _c(ws, row, 2, total,     bg=bg, align="center")
        _c(ws, row, 3, not_iss if not_iss > 0 else "",   bg=LT_RED if not_iss > 0 else bg, align="center", bold=not_iss > 0, fg="C00000" if not_iss > 0 else BLACK)
        _c(ws, row, 4, hold if hold > 0 else "",          bg=AMBER  if hold > 0 else bg,   align="center", bold=hold > 0,   fg="7F4000" if hold > 0 else BLACK)

        pu_cell = _c(ws, row, 5, pct_under, bg=pu_bg, align="center", bold=True, fg=pu_fg)
        pu_cell.font = Font(name="Arial", bold=True, color=pu_fg, size=9)

        pc_cell = _c(ws, row, 6, pct_issued, bg=pc_bg, align="center", bold=True, fg=pc_fg)
        pc_cell.font = Font(name="Arial", bold=True, color=pc_fg, size=9)

        ws.row_dimensions[row].height = 14
        row += 1

    # ── Section 2 TOTAL row ───────────────────────────────────────────────
    all_s = df2[status_col].astype(str).str.strip().str.upper() if status_col in df2.columns else pd.Series(dtype=str)
    grand_total    = len(df2)
    grand_not_iss  = (all_s == "NOT ISSUED").sum()
    grand_hold     = all_s.str.contains("CORRECTION|HOLD", na=False).sum()
    grand_cpy      = (all_s == "ISSUED").sum()
    grand_pct_u    = f"{grand_not_iss / grand_total * 100:.1f}%" if grand_total > 0 else "0.0%"
    grand_pct_c    = f"{grand_cpy / grand_total * 100:.1f}%" if grand_total > 0 else "0.0%"

    _c(ws, row, 1, "TOTAL",        bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    _c(ws, row, 2, grand_total,    bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    _c(ws, row, 3, grand_not_iss,  bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    _c(ws, row, 4, grand_hold,     bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    _c(ws, row, 5, grand_pct_u,    bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    _c(ws, row, 6, grand_pct_c,    bg=DK_ORANGE, bold=True, fg=WHITE, align="center")
    ws.row_dimensions[row].height = 18


def _build_comp5_detail_tab(wb, tab_name, tab_color, df, alt_bg, report_date_str):
    """
    Detail tab with 13 columns as per spec.
    Column M (PCON TR Issue Status) colour-coded.
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A3"

    n_cols = 13

    # Row 1 — title banner
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws.cell(1, 1, f"{tab_name.upper()}   |   COMP5 PROJECT   |   {report_date_str}")
    t.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=tab_color)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — column headers
    col_hdrs = [
        "#", "Client Doc No. (with Rev)", "Saipem Number", "Rev",
        "Title / Description", "Disc.", "Discipline Name",
        "Issue Type", "Date Issued", "Transmittal Reference",
        "Issued by DC", "PCON TR", "PCON TR Issue Status",
    ]
    col_widths = [5, 32, 18, 6, 42, 7, 22, 14, 14, 22, 16, 16, 22]
    for ci, (h, w) in enumerate(zip(col_hdrs, col_widths), 1):
        _h(ws, 2, ci, h, bg=tab_color)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 30
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"

    if df.empty:
        ws.merge_cells(f"A3:{get_column_letter(n_cols)}3")
        ws.cell(3, 1, "No records for this category").font = Font(italic=True, color=GREY595, size=9)
        return

    # Sort: Discipline → Date Issued → Client Doc No
    sort_cols = []
    if "Discipline" in df.columns:      sort_cols.append("Discipline")
    if "Date Issued" in df.columns:     sort_cols.append("Date Issued")
    client_col = next((c for c in df.columns if "CLIENT" in c.upper() and "DOC" in c.upper()), None)
    if client_col:                      sort_cols.append(client_col)
    if sort_cols:
        df = df.sort_values(sort_cols, na_position="last")

    # Map source columns
    def _g(row_data, *keys):
        for k in keys:
            if k in row_data.index:
                return _fmt(row_data[k])
        return ""

    status_col = "PCON- TR Issue Status"
    STATUS_COLORS = {
        "ISSUED":                   (LT_GREEN, DK_GREEN),
        "NOT ISSUED":               (LT_RED,   "C00000"),
        "UNDER CORRECTION/HOLD":    (AMBER,    "7F4000"),
    }

    for ri, (_, row_data) in enumerate(df.iterrows(), 3):
        bg = alt_bg if ri % 2 == 0 else WHITE
        disc_code = str(row_data.get("Discipline","")).strip().upper()
        disc_name = DISC_MAP.get(disc_code, disc_code)
        status    = str(row_data.get(status_col,"")).strip()
        status_up = status.upper()
        if status_up in STATUS_COLORS:
            sbg, sfg = STATUS_COLORS[status_up]
        elif "CORRECTION" in status_up or "HOLD" in status_up:
            sbg, sfg = STATUS_COLORS["UNDER CORRECTION/HOLD"]
        else:
            sbg, sfg = (AMBER, "7F6000")

        _c(ws, ri,  1, ri-2, bg=bg, align="center")
        # Client doc no with rev
        client_no = _g(row_data, "CLIENT DOCUMENT NO. with REV", "CLIENTDOCUMENTNO.")
        _c(ws, ri,  2, client_no, bg=bg)
        _c(ws, ri,  3, _g(row_data, "Saipem Number","SAIPEM NUMBER","Saipem Doc No"), bg=bg)
        _c(ws, ri,  4, _g(row_data, "Rev"), bg=bg, align="center")
        _c(ws, ri,  5, _g(row_data, "TITLE / DESCRIPTION","Title","TITLE"), bg=bg)
        _c(ws, ri,  6, disc_code, bg=bg, align="center")
        _c(ws, ri,  7, disc_name, bg=bg)
        _c(ws, ri,  8, _g(row_data, "Issuing Description"), bg=bg, align="center")
        _c(ws, ri,  9, _g(row_data, "Date Issued"), bg=bg, align="center")
        _c(ws, ri, 10, _g(row_data, "Transmittal Reference"), bg=bg)
        _c(ws, ri, 11, _g(row_data, "Issued by DC"), bg=bg)
        _c(ws, ri, 12, _g(row_data, "PCON TR"), bg=bg)
        sc = _c(ws, ri, 13, status, bg=sbg, align="center", bold=True)
        sc.font = Font(name="Arial", bold=True, color=sfg, size=9)
        ws.row_dimensions[ri].height = 14


def generate_comp5(raw: bytes) -> dict:
    """
    Generate Weekly COMP5 Issued Documents Report.
    Implements COMP5_Report_Instructions.docx exactly.
    6 tabs: Summary | Date-Wise Breakdown | Open-Issued to PCON |
            Under Process at PCON | Pending with Engineering | Issued to CPY
    """
    report_date_str = datetime.today().strftime("%d %B %Y")
    date_str        = datetime.today().strftime("%d%b%Y").upper()

    df_raw = _read_comp5(raw)        # LP → SH already done here
    df     = _dedup_comp5(df_raw)    # latest rev per doc
    df_all, df_issued, df_not_issued, df_hold = _split_comp5(df)

    wb = Workbook(); wb.remove(wb.active)

    _build_comp5_summary(wb, df_all, df_issued, df_not_issued, df_hold, report_date_str)
    _build_datewise_tab(wb, df_all, report_date_str)
    _build_comp5_detail_tab(wb, "Open - Issued to PCON",      "2E75B6",  df_all,        LT_BLUE,   report_date_str)
    _build_comp5_detail_tab(wb, "Under Process at PCON",      DK_ORANGE, df_not_issued, LT_ORANGE, report_date_str)
    _build_comp5_detail_tab(wb, "Pending with Engineering",   PURPLE,    df_hold,       LT_PURPLE, report_date_str)
    _build_comp5_detail_tab(wb, "Issued to CPY",              DK_GREEN,  df_issued,     LT_GREEN,  report_date_str)

    summary = {
        "week":       report_date_str,
        "count_week": len(df_all),
        "count_all":  len(df_all),
        "issued_cpy": len(df_issued),
        "under_proc": len(df_not_issued),
        "pending_eng":len(df_hold),
    }
    filename = f"COMP5_Weekly_IssuedDocs_{date_str}.xlsx"
    return {"bytes": _wb_bytes(wb), "filename": filename, "summary": summary}
